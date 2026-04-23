import openpyxl
import re # For regular expressions (parsing notes, SKUs, prices)
import time
import os
import threading # For periodic saving
import itertools # For finding product combinations
from decimal import Decimal, InvalidOperation # For accurate price calculations

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, StaleElementReferenceException, UnexpectedAlertPresentException
)
from webdriver_manager.chrome import ChromeDriverManager

# --- Configuration ---
# Excel path is overridden by CLI arg --excel or by env var EXCEL_FILE_PATH; defaults to ./data/Return.xlsx
EXCEL_FILE_PATH = os.environ.get('EXCEL_FILE_PATH', os.path.abspath(os.path.join(os.getcwd(), 'data', 'Return.xlsx')))
SHEET_NAME = "Return"
ADMIN_URL = "https://sanveesbytony.com/admin"
# IMPORTANT: Avoid hardcoding credentials. Use environment variables or getpass.
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', "sanenime@gmail.com")  # can be overridden via environment
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', "25694470910")  # can be overridden via environment

INVOICE_DETAIL_BASE_URL = "https://sanveesbytony.com/order/" # Base URL for direct invoice detail access
PRODUCTS_PAGE_URL = "https://sanveesbytony.com/admin/shop/products" # URL for product search page

# --- Column Names (Case-sensitive, ensure they match Excel) ---
COL_INVOICE = "Invoice"
COL_AMOUNT_STATUS = "Amount Status"
COL_NOTE = "Note"
COL_BUYING_PRICE = "Buying Price"
COL_SELLING_PRICE = "Selling Price"
COL_QUANTITY = "Quantity" # Added Quantity column constant

# --- Selectors (Keep these updated if the website changes!) ---
# Login
SELECTOR_EMAIL_INPUT = "#data\\.email"
SELECTOR_PASSWORD_INPUT = "#data\\.password"
SELECTOR_LOGIN_BUTTON = "#form > div.fi-form-actions > div > button"
# Navigation
SELECTOR_INVOICE_MENU_ITEM = "body > div.fi-layout > aside > nav > ul > li:nth-child(2) > ul > li:nth-child(2) > a" # Adjusted for potential structure changes, double-check
SELECTOR_PRODUCT_MENU_ITEM = "body > div.fi-layout > aside > nav > ul > li.fi-sidebar-group.fi-active > ul > li:nth-child(1) > a" # May need refinement based on actual state
# Invoice Search & Details
SELECTOR_INVOICE_SEARCH_INPUT = "#input-1"
SELECTOR_INVOICE_VIEW_LINK = "tbody > tr.fi-ta-row > td.fi-ta-actions-cell > div > div > a" # Simplified, assuming first row result is correct
SELECTOR_INVOICE_DETAIL_PRODUCT_TABLE = 'table.border-bottom > tbody.strong' # Invoice detail page product table body
SELECTOR_INVOICE_DETAIL_PRODUCT_ROWS = f'{SELECTOR_INVOICE_DETAIL_PRODUCT_TABLE} > tr' # Rows within the product table
# Product Search & Edit
SELECTOR_PRODUCT_SEARCH_INPUT = 'input[placeholder="Search"]' # New selector using placeholder
SEARCH_RESULTS_TABLE_BODY_SELECTOR = 'table tbody' # Selector for the table body containing search results
# Buying Price
XPATH_BUYING_PRICE_INPUT = '//*[@id="data.cost"]'
XPATH_NOT_FOUND_INDICATOR = "//div[contains(text(), 'Not Found')]" # XPATH for 404 indicator

# --- Other Settings ---
WAIT_TIMEOUT = 20 # Seconds to wait for elements
SAVE_INTERVAL = 60 # Seconds between automatic saves
MAX_RETRIES = 2 # Max retries for certain Selenium actions

# --- Global Variables ---
driver = None
workbook = None
sheet = None
save_lock = threading.Lock()
stop_saving_event = threading.Event()
last_processed_row_index = 1 # Keep track for saving progress

# --- Helper Functions ---

def clean_price(price_str):
    """Removes currency symbols, commas, and converts to Decimal."""
    if price_str is None:
        return Decimal(0)
    # Remove "Taka", commas, and whitespace
    cleaned = re.sub(r'[^\d.]', '', str(price_str).replace(',', ''))
    try:
        return Decimal(cleaned) if cleaned else Decimal(0)
    except InvalidOperation:
        print(f"Warning: Could not convert price '{price_str}' to Decimal. Using 0.")
        return Decimal(0)

def extract_sku(sku_text):
    """Extracts SKU, handling different formats."""
    if sku_text is None:
        return None
    sku_text = sku_text.strip()
    # Case 1: "SKU: XXX (YYY)" -> Extract YYY
    match_bracket = re.search(r'\(([^)]+)\)', sku_text)
    if match_bracket:
        return match_bracket.group(1).strip()
    # Case 2: "SKU: ZZZ" -> Extract ZZZ
    match_simple = re.search(r'SKU:\s*(.*)', sku_text, re.IGNORECASE)
    if match_simple:
        return match_simple.group(1).strip()
    # Fallback: Return the text if no pattern matches (might be just the SKU)
    print(f"Warning: Could not parse SKU format for '{sku_text}'. Using raw value.")
    return sku_text

def parse_note(note_text):
    """Parses the note to determine the return type and target amount."""
    if not isinstance(note_text, str):
        return "unknown", None # Handle non-string notes

    note_text = note_text.strip()

    if note_text == "Full Amount Returned":
        return "full_return", None

    # Regex for "Amount has been changed from "X" to "Y""
    match = re.match(r'Amount has been changed from "(\d+)" to "(\d+)"', note_text, re.IGNORECASE)
    if match:
        to_amount = int(match.group(2))
        if 0 <= to_amount <= 150:
            return "full_return_charge_only", None # Treat as full return
        else:
            # Partial return, target amount comes from 'Amount Status' column
            return "partial_return_amount_in_status", None

    return "unknown", None # Default if no pattern matches

def safe_find_element(by, value, wait_time=WAIT_TIMEOUT, retries=MAX_RETRIES):
    """Finds an element with explicit wait and retries."""
    for attempt in range(retries):
        try:
            wait = WebDriverWait(driver, wait_time)
            element = wait.until(EC.presence_of_element_located((by, value)))
            # Optional: Check for visibility as well if needed
            # element = wait.until(EC.visibility_of_element_located((by, value)))
            return element
        except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
            print(f"Attempt {attempt + 1}/{retries}: Element ({by}, {value}) not found or stale after {wait_time}s. Retrying...")
            time.sleep(1) # Short pause before retry
            if attempt == retries - 1:
                print(f"Error: Element ({by}, {value}) not found after {retries} attempts. Error: {e}")
                return None # Return None if all retries fail

def safe_find_elements(by, value, wait_time=WAIT_TIMEOUT, retries=MAX_RETRIES):
    """Finds multiple elements with explicit wait and retries."""
    for attempt in range(retries):
        try:
            wait = WebDriverWait(driver, wait_time)
            # Wait for at least one element to be present
            wait.until(EC.presence_of_element_located((by, value)))
            # Then get all matching elements
            elements = driver.find_elements(by, value)
            return elements
        except (TimeoutException, StaleElementReferenceException) as e:
            print(f"Attempt {attempt + 1}/{retries}: Elements ({by}, {value}) not found or stale after {wait_time}s. Retrying...")
            time.sleep(1) # Short pause before retry
            if attempt == retries - 1:
                print(f"Error: Elements ({by}, {value}) not found after {retries} attempts. Error: {e}")
                return [] # Return empty list if all retries fail


def safe_click(element, wait_time=WAIT_TIMEOUT, retries=MAX_RETRIES):
    """Clicks an element with explicit wait for clickability and retries."""
    if not element:
        print("Error: Cannot click a None element.")
        return False
    for attempt in range(retries):
        try:
            wait = WebDriverWait(driver, wait_time)
            clickable_element = wait.until(EC.element_to_be_clickable(element))
            clickable_element.click()
            return True # Click successful
        except (TimeoutException, StaleElementReferenceException, NoSuchElementException) as e:
            print(f"Attempt {attempt + 1}/{retries}: Could not click element. Retrying...")
            time.sleep(1)
            # Re-find the element in case it went stale
            try:
                element = driver.find_element(element.by, element.value)
            except:
                 print(f"Error re-finding element for click retry.")
                 if attempt == retries - 1: return False
                 continue # Try next attempt

            if attempt == retries - 1:
                print(f"Error: Failed to click element after {retries} attempts. Error: {e}")
                return False # Click failed

def safe_send_keys(element, keys, wait_time=WAIT_TIMEOUT):
    """Sends keys to an element with explicit wait."""
    if not element:
        print("Error: Cannot send keys to a None element.")
        return False
    try:
        wait = WebDriverWait(driver, wait_time)
        visible_element = wait.until(EC.visibility_of(element))
        visible_element.clear() # Clear first
        visible_element.send_keys(keys)
        return True
    except (TimeoutException, StaleElementReferenceException, NoSuchElementException) as e:
        print(f"Error: Could not send keys to element. Error: {e}")
        return False

def periodic_saver():
    """Saves the workbook periodically in a separate thread."""
    global workbook
    global last_processed_row_index
    while not stop_saving_event.is_set():
        time.sleep(SAVE_INTERVAL)
        with save_lock: # Acquire lock before accessing/saving workbook
            if workbook:
                try:
                    # Create a backup before saving
                    # backup_path = EXCEL_FILE_PATH + f".backup_{time.strftime('%Y%m%d_%H%M%S')}"
                    workbook.save(EXCEL_FILE_PATH)
                    print(f"--- Workbook automatically saved (processed up to row {last_processed_row_index}) ---")
                except Exception as e:
                    print(f"--- Error during periodic save: {e} ---")
    print("--- Periodic saving thread stopped. ---")

def find_product_combination(products, target_amount):
    """Finds a combination of products (with quantities) matching the target selling price."""
    target_amount = clean_price(target_amount)
    if target_amount <= 0: return [] # No need to search if target is zero or less

    # Create a list representing individual items (quantity > 1 means multiple entries)
    individual_items = []
    for i, p in enumerate(products):
        for _ in range(p['quantity']):
            individual_items.append({'index': i, **p}) # Keep original index

    n = len(individual_items)
    print(f"Searching for combination matching {target_amount} among {n} individual items...")

    # Iterate through possible numbers of returned items (from 1 up to n)
    for k in range(1, n + 1):
        # Generate combinations of k items
        for combo_indices in itertools.combinations(range(n), k):
            current_combo_items = [individual_items[idx] for idx in combo_indices]
            current_sum = sum(item['selling_price'] for item in current_combo_items)

            # Use Decimal comparison
            if abs(current_sum - target_amount) < Decimal('0.01'): # Allow for small floating point differences
                print(f"Found combination matching {target_amount}: {[(item['sku'], item['selling_price']) for item in current_combo_items]}")
                # Consolidate back into original product list format with quantities
                result = []
                processed_indices = set()
                for item in current_combo_items:
                    original_index = item['index']
                    if original_index not in processed_indices:
                        count = sum(1 for ci in current_combo_items if ci['index'] == original_index)
                        result.append({**products[original_index], 'returned_quantity': count}) # Add returned_quantity
                        processed_indices.add(original_index)
                return result

    print(f"Warning: Could not find exact combination matching {target_amount}.")
    return [] # Return empty if no combination found


# --- Main Script Logic ---

def setup_driver():
    """Initializes the Selenium WebDriver."""
    global driver
    print("Setting up Chromium browser...")
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("window-size=1920x1080")
    # Optional: Specify user data dir for persistent sessions (might avoid frequent logins)
    # chrome_options.add_argument("--user-data-dir=/path/to/your/chrome/profile")

    # Add this line for headless mode
    chrome_options.add_argument("--headless=new")

    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.implicitly_wait(5) # Implicit wait as a small fallback
        print("WebDriver setup successful.")
        return True
    except Exception as e:
        print(f"Error setting up WebDriver: {e}")
        return False

def login():
    """Logs into the admin panel."""
    global driver
    if not driver: return False

    print(f"Navigating to login page: {ADMIN_URL}")
    try:
        driver.get(ADMIN_URL)
    except Exception as e:
        print(f"Error navigating to login page: {e}")
        return False

    print("Entering credentials...")
    email_input = safe_find_element(By.CSS_SELECTOR, SELECTOR_EMAIL_INPUT)
    password_input = safe_find_element(By.CSS_SELECTOR, SELECTOR_PASSWORD_INPUT)
    login_button = safe_find_element(By.CSS_SELECTOR, SELECTOR_LOGIN_BUTTON)

    if not email_input or not password_input or not login_button:
        print("Error: Could not find login elements.")
        return False

    if not safe_send_keys(email_input, ADMIN_EMAIL): return False
    if not safe_send_keys(password_input, ADMIN_PASSWORD): return False

    print("Attempting login...")
    if not safe_click(login_button):
        print("Error: Failed to click login button.")
        return False

    # Wait for a known element on the dashboard to confirm login is complete
    # Adjust selector if needed based on observation
    DASHBOARD_READY_SELECTOR = "main > div > section" # Keep current, but be ready to refine
    print("Waiting for dashboard to load and become ready...")
    try:
        wait = WebDriverWait(driver, WAIT_TIMEOUT)
        # Wait for the element to be present and visible
        dashboard_element_ready = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, DASHBOARD_READY_SELECTOR))
        )
        print("Dashboard element found and is visible.")

        # --- Add a small sleep here as a test ---
        time.sleep(3) # Add a small pause (adjust as needed)
        print("Added a short post-login pause.")
        # --- End of temporary sleep ---

        print("Login successful.")
        return True
    except TimeoutException:
        print("Login failed or dashboard did not load/become visible within timeout.")
        # Check for error messages if possible on the login page
        # error_msg = safe_find_element(By.CSS_SELECTOR, ".login-error-message-selector")
        # if error_msg: print(f"Login error message: {error_msg.text}")
        return False
    except Exception as e:
        print(f"An unexpected error occurred during login waiting: {e}")
        return False

def get_invoice_product_details(invoice_id):
    """Navigates directly to the invoice detail page and extracts product details."""
    global driver
    print(f"\n--- Processing Invoice ID: {invoice_id} ---")

    # Construct the direct URL for the invoice detail page
    invoice_detail_url = f"{INVOICE_DETAIL_BASE_URL}{invoice_id}/invoice"
    print(f"Navigating directly to Invoice detail page: {invoice_detail_url}")

    # --- Navigation to Invoice Detail Page with Retry and Re-login ---
    max_nav_retries = 2 # Limit retries
    original_window = driver.current_window_handle # Store original window handle
    nav_outcome = None # Initialize outcome: None=fail, True=success, "not_found"=404

    for nav_attempt in range(max_nav_retries):
        print(f"Attempt {nav_attempt + 1}/{max_nav_retries}: Navigating to {invoice_detail_url}")
        new_window = None # Initialize new_window inside try block

        try:
            # Open the invoice detail page in a new tab by executing JavaScript
            driver.execute_script(f"window.open('{invoice_detail_url}', '_blank');")
            print("Opened invoice detail page in a new tab.")

            # Wait for and Switch to New Tab
            print("Waiting for new invoice detail tab...")
            WebDriverWait(driver, WAIT_TIMEOUT).until(EC.number_of_windows_to_be(2))
            new_window = [window for window in driver.window_handles if window != original_window][0]
            driver.switch_to.window(new_window)
            print("Switched to invoice detail tab.")

            # --- Check if it's a "Not Found" page ---
            print("Checking if page is 'Not Found'...")
            try:
                # Use a very short wait to check for the "Not Found" indicator
                not_found_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, XPATH_NOT_FOUND_INDICATOR))
                )
                # If the element is found, it's a 404 page
                print("'Not Found' indicator found. Invoice does not exist.")
                nav_outcome = "not_found" # Mark as not found
                break # Exit the navigation retry loop

            except TimeoutException:
                # If TimeoutException here, the 'Not Found' element was NOT found within 5s.
                # This is the expected path for a valid invoice page.
                print("'Not Found' indicator not found within 5s. Assuming valid page or other issue.")
                pass # Continue to wait for the product table

            # --- If not a "Not Found" page, wait for the product table to confirm successful load ---
            print("Waiting for product table on invoice detail page...")
            wait = WebDriverWait(driver, WAIT_TIMEOUT - 5) # Give a little less time here since we already waited 5s
            wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR_INVOICE_DETAIL_PRODUCT_TABLE))
            )
            time.sleep(2) # Extra pause for potential dynamic loading

            print("Product table found. Successfully loaded invoice detail page.")
            nav_outcome = True # Mark navigation as successful
            break # Exit the navigation retry loop on success

        except (TimeoutException, UnexpectedAlertPresentException, Exception) as e:
            # Handle other navigation errors or unexpected alerts (like session expired)
            print(f"Error navigating to or loading invoice detail page in new tab: {e}")
            nav_outcome = None # Ensure outcome is marked as failure for other errors

            if isinstance(e, UnexpectedAlertPresentException):
                 print("Detected unexpected alert (likely session expired). Accepting alert and attempting re-login...")
                 try:
                     alert = driver.switch_to.alert
                     alert.accept()
                     print("Alert accepted.")
                     time.sleep(2)
                 except Exception as alert_e:
                     print(f"Error handling alert: {alert_e}")

            # Attempt to re-login if navigation failed or alert occurred
            print("Attempting to re-login...")
            try:
                driver.switch_to.window(original_window)
                # Close the problematic new tab if it was opened
                if new_window and new_window in driver.window_handles:
                     driver.switch_to.window(new_window)
                     driver.close()
                driver.switch_to.window(original_window) # Ensure we are back

            except Exception as switch_e:
                 print(f"Error switching back to original window before re-login: {switch_e}. May affect re-login.")


            if login():
                print("Re-login successful. Retrying navigation...")
                continue # Retry the navigation attempt in the loop
            else:
                print("Re-login failed. Cannot proceed with invoice details.")
                if nav_attempt == max_nav_retries - 1:
                     print(f"Error: Failed to navigate to Invoice detail page after {max_nav_retries} attempts, including re-login attempts.")
                     return None # Indicate failure after max retries


    # --- Check the navigation outcome after the retry loop ---
    if nav_outcome == "not_found":
        # Handled the 'Not Found' case explicitly
        print(f"Invoice {invoice_id} does not exist. Skipping row.")
        # Ensure the new tab is closed and we are back on the original window
        try:
             if new_window and new_window in driver.window_handles:
                  driver.switch_to.window(new_window)
                  driver.close()
             driver.switch_to.window(original_window)
        except Exception as cleanup_e:
             print(f"Error during cleanup after 'not_found': {cleanup_e}")
        return "not_found" # Return special value

    elif nav_outcome is not True:
        # Navigation failed for other reasons (max retries exhausted or unhandled error)
        print(f"Failed to navigate to invoice detail page for {invoice_id} after retries or due to error.")
         # Attempt cleanup if a new window was potentially opened and not closed
        try:
             if new_window and new_window in driver.window_handles:
                  driver.switch_to.window(new_window)
                  driver.close()
             driver.switch_to.window(original_window)
        except Exception as cleanup_e:
             print(f"Error during cleanup after navigation failure: {cleanup_e}")
        return None # Indicate a general failure


    # --- If navigation was successful (nav_outcome is True), proceed with data extraction ---
    # 5. Extract Product Data from Invoice Detail Page
    products = []
    print("Extracting product details from invoice page...")
    try:
        product_rows = safe_find_elements(By.CSS_SELECTOR, SELECTOR_INVOICE_DETAIL_PRODUCT_ROWS)
        if not product_rows:
             print("Warning: No product rows found in the table.")
             # If no product rows but page loaded, might be an empty invoice or different format
             # Decide how to handle this - return empty list or error?
             # For now, we'll return an empty list and let the main loop handle it.

        for row in product_rows:
            try:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) < 5:
                    print(f"Warning: Skipping row, unexpected number of cells ({len(cells)}).")
                    continue

                sku_element = cells[1].find_element(By.TAG_NAME, 'small')
                sku_raw = sku_element.text
                sku = extract_sku(sku_raw)

                quantity = int(cells[2].text.strip())

                selling_price_str = cells[3].text
                selling_price = clean_price(selling_price_str)

                if sku and selling_price >= 0 and quantity > 0: # Allow selling_price to be 0
                    products.append({
                        "sku": sku,
                        "selling_price": selling_price,
                        "quantity": quantity
                    })
                    print(f"  - Extracted: SKU={sku}, Qty={quantity}, SellPrice={selling_price}")
                else:
                     print(f"Warning: Skipping product due to missing or invalid data (SKU: {sku}, Price: {selling_price}, Qty: {quantity})")

            except (NoSuchElementException, ValueError, IndexError, StaleElementReferenceException) as e:
                print(f"Error extracting data from a product row: {e}")
                continue

    except Exception as e:
        print(f"Error finding product rows or during extraction loop: {e}")
        # Ensure cleanup even if extraction fails
        try:
            if driver.current_window_handle != original_window and "invoice" in driver.current_url.lower():
                 driver.close() # Close the current (invoice detail) tab
            driver.switch_to.window(original_window) # Switch back to the original window
        except Exception as cleanup_e:
             print(f"Error during cleanup after extraction failure: {cleanup_e}")

        return None # Indicate extraction failure


    # 6. Close Tab and Switch Back
    print("Closing invoice detail tab and switching back...")
    try:
        # Ensure we are on the invoice detail tab before closing
        # Use a list comprehension to check if any window handle other than the original exists
        other_windows = [w for w in driver.window_handles if w != original_window]
        if other_windows:
            # Assuming the one we want to close is the first of the others
             driver.switch_to.window(other_windows[0])
             if "invoice" in driver.current_url.lower(): # Double check we are on an invoice page
                  driver.close()
             driver.switch_to.window(original_window) # Switch back to the original window
             print("Closed invoice detail tab and switched back.")
        else:
             print("No other window found to close. Already on main window?")
             # If we are already on the main window, try navigating to a known page to be safe
             try:
                  print("Attempting to navigate main window to orders page for state reset...")
                  driver.get(PRODUCTS_PAGE_URL) # Navigate to a known page (e.g., products page)
                  WebDriverWait(driver, WAIT_TIMEOUT).until(EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR_PRODUCT_SEARCH_INPUT)))
                  print("Main window navigation for state reset successful.")
             except Exception as nav_e:
                  print(f"Warning: Could not navigate main window to reset state: {nav_e}")


    except Exception as e:
         print(f"Error closing tab or switching back: {e}")
         # Attempt to recover by navigating the main window back to a known state
         try:
              print("Attempting to recover by navigating main window to products page...")
              # Make sure we are on the main window handle before navigating
              driver.switch_to.window(original_window)
              driver.get(PRODUCTS_PAGE_URL) # Navigate to a known page (e.g., products page)
              # Optionally wait for an element on this page to confirm recovery
              WebDriverWait(driver, WAIT_TIMEOUT).until(EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR_PRODUCT_SEARCH_INPUT)))
              print("Recovery navigation successful.")
         except Exception as recovery_e:
              print(f"Critical Error: Could not recover after failing to close tab or switch back: {recovery_e}")
              # At this point, the driver state might be unpredictable. Consider quitting.
              # driver.quit()
              # exit("Critical error, driver state compromised.")


    return products

def get_product_buying_price(sku):
    """Searches for a product by SKU, finds its edit link URL by pattern, navigates there, and returns its buying price."""
    global driver
    print(f"  Fetching buying price for SKU: {sku}")

    # --- Navigation to Product Search Page with Retry and Re-login ---
    max_nav_retries = 2 # Limit retries

    for nav_attempt in range(max_nav_retries):
        print(f"Attempt {nav_attempt + 1}/{max_nav_retries}: Navigating directly to Product search page: {PRODUCTS_PAGE_URL}")
        try:
            driver.get(PRODUCTS_PAGE_URL)
            # Wait for the search input to be present to confirm successful navigation/login
            WebDriverWait(driver, WAIT_TIMEOUT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR_PRODUCT_SEARCH_INPUT))
            )
            print("  Successfully navigated to Product search page.")
            break # Exit navigation retry loop on success
        except (TimeoutException, UnexpectedAlertPresentException, Exception) as e:
            print(f"  Error navigating directly to product search page: {e}")

            # Handle the "This page has expired" alert specifically
            if isinstance(e, UnexpectedAlertPresentException):
                 print("  Detected unexpected alert (likely session expired). Accepting alert and attempting re-login...")
                 try:
                     alert = driver.switch_to.alert
                     alert.accept() # Accept the alert (usually refreshes the page)
                     print("  Alert accepted.")
                     time.sleep(2) # Give page a moment to refresh/load
                 except Exception as alert_e:
                     print(f"  Error handling alert: {alert_e}")

            # Attempt to re-login if navigation failed or alert occurred
            print("  Attempting to re-login...")
            if login():
                 print("  Re-login successful. Retrying navigation...")
                 continue # Retry the navigation attempt
            else:
                 print("  Re-login failed. Cannot proceed with navigation.")
                 if nav_attempt == max_nav_retries - 1:
                      print(f"  Error: Failed to navigate to Product search page after {max_nav_retries} attempts, including re-login attempts.")
                      return None # Indicate failure after max retries

    # If navigation failed after retries, return None
    if nav_attempt == max_nav_retries - 1 and (not driver.current_url.startswith(PRODUCTS_PAGE_URL) or not safe_find_element(By.CSS_SELECTOR, SELECTOR_PRODUCT_SEARCH_INPUT)):
         print("  Failed to reach product search page after retries.")
         return None


    # 2. Search for SKU
    print(f"  Searching for SKU {sku}...")
    search_input = safe_find_element(By.CSS_SELECTOR, SELECTOR_PRODUCT_SEARCH_INPUT)
    if not search_input:
         print(f"  Error: Product search input not found after navigation for SKU {sku}.")
         # Attempt to return to a known state in case of element not found
         try: driver.get(PRODUCTS_PAGE_URL) # Try navigating back to search page
         except: pass
         return None

    if not safe_send_keys(search_input, sku):
         print(f"  Error: Could not input SKU {sku}.")
         # Attempt to return to a known state
         try: driver.get(PRODUCTS_PAGE_URL)
         except: pass
         return None

    # IMPORTANT: Wait for search results to load before looking for links
    print("  Waiting for search results to load...")
    try:
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, SEARCH_RESULTS_TABLE_BODY_SELECTOR))
        )
        print("  Search results table body found.")
        time.sleep(2) # Allow potential dynamic content
    except TimeoutException:
        print(f"  Warning: Search results table body did not appear within {WAIT_TIMEOUT}s.")
        # Decide how to handle this - if no table, probably no results.
        # Proceed to look for links, but expect none.
        pass

    # 3. Find the Product Edit Link by URL Pattern
    print("  Looking for product edit link by URL pattern...")
    product_edit_url = None
    
    # --- Try finding link by XPath first as requested ---
    xpath_try_1 = "/html/body/div[1]/div[1]/main/div/section/div/div/div/div/div/div[4]/table/tbody/tr/td[12]/div/div/a[2]/span"
    xpath_try_2 = "/html/body/div[1]/div[1]/main/div/section/div/div/div/div/div/div[4]/table/tbody/tr[1]/td[12]/div/div/a[2]/span"
    
    # Note: The user provided XPaths to a span inside an anchor. We need the anchor's href.
    # We will try to find the element and then get the parent anchor or check if the element itself is clickable/has href context.
    # Looking at the structure ".../a[2]/span", the link is the parent of the span.
    
    found_link_element = None
    
    # Try XPath 1
    try:
        element = driver.find_element(By.XPATH, xpath_try_1)
        # The xpath points to a span, we need the parent 'a' tag
        found_link_element = element.find_element(By.XPATH, "./..") 
        print(f"  Found product edit link using XPath 1.")
    except NoSuchElementException:
        # Try XPath 2
        try:
            element = driver.find_element(By.XPATH, xpath_try_2)
            found_link_element = element.find_element(By.XPATH, "./..")
            print(f"  Found product edit link using XPath 2.")
        except NoSuchElementException:
            print("  Could not find product edit link using provided XPaths. Falling back to URL pattern search.")

    if found_link_element:
        product_edit_url = found_link_element.get_attribute('href')
    
    # Fallback to original logic if XPath method failed
    if not product_edit_url:
        links_container_element = safe_find_element(By.CSS_SELECTOR, SEARCH_RESULTS_TABLE_BODY_SELECTOR)
        if links_container_element:
            link_elements = links_container_element.find_elements(By.TAG_NAME, 'a')
            print(f"  Found {len(link_elements)} links in the results container.")
        else:
            print("  Could not find results container, searching all links on the page.")
            link_elements = driver.find_elements(By.TAG_NAME, 'a')
            print(f"  Found {len(link_elements)} links on the page.")


        url_pattern = re.compile(r"https://sanveesbytony\.com/admin/shop/products/\d+/edit")

        for link_element in link_elements:
            try:
                href = link_element.get_attribute('href')
                if href and url_pattern.match(href):
                    product_edit_url = href
                    print(f"  Found matching product edit URL: {product_edit_url}")
                    break
            except StaleElementReferenceException:
                print("  Warning: Stale element reference while checking links. Skipping element.")
                continue
            except Exception as e:
                 print(f"  Error checking link href: {e}. Skipping element.")
                 continue

    if not product_edit_url:
        print(f"  Error: Could not find product edit link with matching URL pattern for SKU {sku}. Likely no search results or wrong page.")
        # Attempt to return to a known state
        try: driver.get(PRODUCTS_PAGE_URL)
        except: pass
        return None

    # 4. Navigate Directly to the Product Edit Page URL
    print(f"  Navigating directly to product edit page: {product_edit_url}")
    try:
        driver.get(product_edit_url)
        WebDriverWait(driver, WAIT_TIMEOUT).until(
             EC.presence_of_element_located((By.XPATH, XPATH_BUYING_PRICE_INPUT))
        )
        print("  Successfully navigated to product edit page.")
    except Exception as e:
        print(f"  Error navigating to product edit page: {e}")
         # Attempt to return to a known state
        try: driver.get(PRODUCTS_PAGE_URL)
        except: pass
        return None

    # 5. Extract Buying Price from the Edit Page
    print("  Extracting buying price...")
    buying_price_input = safe_find_element(By.XPATH, XPATH_BUYING_PRICE_INPUT)
    if not buying_price_input:
        print(f"  Error: Buying price input element ({XPATH_BUYING_PRICE_INPUT}) not found on edit page for SKU {sku}.")
        # Attempt to return to a known state
        try: driver.get(PRODUCTS_PAGE_URL)
        except: pass
        return None

    try:
        buying_price_str = buying_price_input.get_attribute('value')
        buying_price = clean_price(buying_price_str)
        print(f"  Found Buying Price for {sku}: {buying_price}")
         # Attempt to return to a known state after fetching price
        try: driver.get(PRODUCTS_PAGE_URL)
        except: pass
        return buying_price
    except Exception as e:
        print(f"  Error reading buying price value for SKU {sku}: {e}")
         # Attempt to return to a known state
        try: driver.get(PRODUCTS_PAGE_URL)
        except: pass
        return None

def update_excel(row_index, buying_total, selling_total, returned_quantity, col_map):
    """Updates the Buying Price, Selling Price, and Quantity columns for a specific row."""
    global sheet
    try:
        # Use Decimal for currency, format as string for Excel if needed
        buy_val = str(buying_total) if buying_total is not None else "Error/NotFound"
        sell_val = str(selling_total) if selling_total is not None else "Error/NotFound"
        qty_val = str(returned_quantity) if returned_quantity is not None else "Error/NotFound" # Handle quantity value

        if buying_total == "none" and selling_total == "none" and returned_quantity == "none":
             buy_val = "none"
             sell_val = "none"
             qty_val = "none" # Also set quantity to 'none' in this case

        sheet.cell(row=row_index, column=col_map[COL_BUYING_PRICE], value=buy_val)
        sheet.cell(row=row_index, column=col_map[COL_SELLING_PRICE], value=sell_val)
        sheet.cell(row=row_index, column=col_map[COL_QUANTITY], value=qty_val) # Update the Quantity column
        print(f"Updated Excel Row {row_index}: Buying={buy_val}, Selling={sell_val}, Quantity={qty_val}")
        return True
    except Exception as e:
        print(f"Error updating Excel row {row_index}: {e}")
        return False

# Treat empty/placeholder/NaN cells as empty
def cell_is_empty(val):
    if val is None:
        return True
    # handle float NaN without importing math
    try:
        if isinstance(val, float) and val != val:
            return True
    except Exception:
        pass
    if isinstance(val, str):
        low = val.strip().lower()
        if low in ("", "none", "nan"):
            return True
        if "error" in low or "notfound" in low:
            return True
    return False

# --- Main Execution ---
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Post-process Return.xlsx to add Buying/Selling/Quantity")
    parser.add_argument('--excel', '-e', help='Path to Excel file to process')
    args = parser.parse_args()

    if args.excel:
        EXCEL_FILE_PATH = args.excel

    # --- Initial Setup ---
    if not setup_driver():
        exit("Failed to initialize WebDriver. Exiting.")

    if not login():
        driver.quit()
        exit("Login failed. Exiting.")

    # --- Load Excel ---
    try:
        print(f"Loading Excel file: {EXCEL_FILE_PATH}")
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook[SHEET_NAME]
        print(f"Loaded sheet: {SHEET_NAME}")

        # Find column indices
        header = [cell.value for cell in sheet[1]]
        column_map = {name: i + 1 for i, name in enumerate(header)}

        # Validate required columns exist
        required_cols = [COL_INVOICE, COL_AMOUNT_STATUS, COL_NOTE, COL_BUYING_PRICE, COL_SELLING_PRICE, COL_QUANTITY] # Added COL_QUANTITY
        if not all(col in column_map for col in required_cols):
            missing = [col for col in required_cols if col not in column_map]
            print(f"Error: Missing required columns in Excel sheet: {missing}. Please add them to your Excel file.")
            driver.quit()
            exit()

    except FileNotFoundError:
        print(f"Error: Excel file not found at {EXCEL_FILE_PATH}")
        driver.quit()
        exit()
    except KeyError:
        print(f"Error: Sheet named '{SHEET_NAME}' not found in the Excel file.")
        driver.quit()
        exit()
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        driver.quit()
        exit()

    # --- Start Periodic Saving ---
    save_thread = threading.Thread(target=periodic_saver, daemon=True)
    save_thread.start()
    print("Periodic saving thread started.")

    # --- Process Rows ---
    # Iterate from row 2 (skip header)
    for row_index in range(2, sheet.max_row + 1):
        last_processed_row_index = row_index # Update for saver thread
        with save_lock: # Ensure main thread doesn't modify sheet structure while saving
             # Check if already processed (optional, based on whether BP/SP cols are empty or QTY is filled)
            current_bp = sheet.cell(row=row_index, column=column_map[COL_BUYING_PRICE]).value
            current_sp = sheet.cell(row=row_index, column=column_map[COL_SELLING_PRICE]).value
            current_qty = sheet.cell(row=row_index, column=column_map[COL_QUANTITY]).value # Get current quantity

            # Basic check: skip only if all three cells are actually filled (not empty/NaN/placeholder)
            if (not cell_is_empty(current_bp)) and (not cell_is_empty(current_sp)) and (not cell_is_empty(current_qty)):
                 print(f"Row {row_index}: Skipping, Buying Price, Selling Price, and Quantity already filled ('{current_bp}', '{current_sp}', '{current_qty}').")
                 continue

            # Get data for the current row
            invoice_id = sheet.cell(row=row_index, column=column_map[COL_INVOICE]).value
            amount_status_str = sheet.cell(row=row_index, column=column_map[COL_AMOUNT_STATUS]).value
            note = sheet.cell(row=row_index, column=column_map[COL_NOTE]).value


        # Basic validation
        if not invoice_id:
            print(f"Row {row_index}: Skipping, missing Invoice ID.")
            with save_lock: update_excel(row_index, "Error: No ID", "Error: No ID", "Error: No ID", column_map)
            continue

        print(f"\n>>> Processing Row: {row_index}, Invoice: {invoice_id}, Note: '{note}', Amount Status: '{amount_status_str}'")

        # 1. Get Product Details from Invoice Page
        invoice_products = get_invoice_product_details(invoice_id)

        if invoice_products == "not_found":
            print(f"Invoice {invoice_id} not found. Updating Excel as 'none'.")
            with save_lock: update_excel(row_index, "none", "none", "none", column_map)
            continue # Move to next row
        elif invoice_products is None:
            print(f"Failed to get product details for invoice {invoice_id}. Skipping row.")
            # Optionally mark the row as errored in Excel
            with save_lock: update_excel(row_index, "Error: Fetch", "Error: Fetch", "Error: Fetch", column_map)
            continue
        elif not invoice_products:
             print(f"No products found listed on invoice {invoice_id}. Updating Excel as 0.")
             with save_lock: update_excel(row_index, 0, 0, 0, column_map) # Pass 0 for quantity
             continue


        # 2. Determine Which Products to Process
        return_type, _ = parse_note(note)
        products_to_process = []

        if return_type == "full_return" or return_type == "full_return_charge_only":
            print("Processing as Full Return (or charge only).")
            # Mark all products with their original quantity as 'returned'
            products_to_process = [{**p, 'returned_quantity': p['quantity']} for p in invoice_products]

        elif return_type == "partial_return_amount_in_status":
            print("Processing as Partial Return based on Amount Status.")
            target_return_amount = clean_price(amount_status_str)
            if target_return_amount > 0:
                 products_to_process = find_product_combination(invoice_products, target_return_amount)
                 if not products_to_process:
                     print(f"Warning: Could not find product combination for target amount {target_return_amount}. Marking row as Error.")
                     with save_lock: update_excel(row_index, "Error: Combo", "Error: Combo", "Error: Combo", column_map)
                     continue # Skip buying price lookup if combo fails
            else:
                 print("Warning: Partial return indicated, but Amount Status is zero or invalid. Skipping.")
                 with save_lock: update_excel(row_index, "Error: Status 0", "Error: Status 0", "Error: Status 0", column_map)
                 continue

        else: # Unknown note type
            print(f"Unknown or unhandled Note type: '{note}'. Skipping row processing.")
            # Optionally mark as needing manual review
            with save_lock: update_excel(row_index, "Manual Review", "Manual Review", "Manual Review", column_map)
            continue

        # 3. Fetch Buying Prices and Calculate Totals
        total_buying_price = Decimal(0)
        total_selling_price = Decimal(0)
        total_returned_quantity = 0 # Initialize total returned quantity
        buy_price_fetch_failed = False

        if not products_to_process:
             print("No products determined for processing based on criteria.")
             # Update Excel as 0,0,0 perhaps? Or based on specific logic?
             with save_lock: update_excel(row_index, 0, 0, 0, column_map) # Pass 0 for quantity
             continue

        print(f"Fetching buying prices for {len(products_to_process)} determined product type(s)...")
        for product in products_to_process:
            sku = product["sku"]
            returned_qty = product["returned_quantity"]
            selling_price_per_item = product["selling_price"]

            buying_price_per_item = get_product_buying_price(sku)

            if buying_price_per_item is None:
                print(f"  -> Failed to get buying price for SKU {sku}. Aborting calculation for this row.")
                buy_price_fetch_failed = True
                # Decide how to handle quantity here - maybe set to 0 or 'Error'?
                # If buying price fetch fails, we mark all outputs as error for this row
                total_returned_quantity = "Error"
                total_buying_price = "Error"
                total_selling_price = "Error"
                break # Stop processing this row if any buying price fails

            # Add to totals based on the *returned* quantity
            # Ensure total_buying_price and total_selling_price are Decimal before addition
            if not isinstance(total_buying_price, Decimal): total_buying_price = Decimal(0)
            if not isinstance(total_selling_price, Decimal): total_selling_price = Decimal(0)

            total_buying_price += buying_price_per_item * returned_qty
            total_selling_price += selling_price_per_item * returned_qty

            # Ensure total_returned_quantity is an integer before addition
            if not isinstance(total_returned_quantity, int): total_returned_quantity = 0
            total_returned_quantity += returned_qty

            print(f"  -> SKU {sku}: Returned Qty={returned_qty}, Buy={buying_price_per_item}, Sell={selling_price_per_item} | Running Totals: Buy={total_buying_price}, Sell={total_selling_price}, Qty={total_returned_quantity}")

        # 4. Update Excel with Calculated Totals and Quantity
        with save_lock:
             if buy_price_fetch_failed:
                 # Pass the error state for quantity as well
                 update_excel(row_index, total_buying_price, total_selling_price, total_returned_quantity, column_map)
             # The "not_found" case for products_to_process is handled above, before fetching buying prices.
             # If we reach here, products_to_process was either a list (potentially empty) or 'none'.
             # If it was an empty list, the totals will be 0 and handled correctly by the else block.
             else:
                 # Pass the calculated total_returned_quantity
                 update_excel(row_index, total_buying_price, total_selling_price, total_returned_quantity, column_map)


    # --- Final Save and Cleanup ---
    print("\nProcessing complete. Performing final save...")
    stop_saving_event.set() # Signal the saver thread to stop
    save_thread.join(timeout=5) # Wait briefly for the thread to finish

    with save_lock: # Final save by the main thread
        if workbook:
            try:
                workbook.save(EXCEL_FILE_PATH)
                print(f"Final workbook saved successfully to {EXCEL_FILE_PATH}")
            except Exception as e:
                print(f"Error during final save: {e}")

    if driver:
        driver.quit()
        print("Browser closed.")

    print("Script finished.")